#!/usr/bin/env python3
"""Build Excel workbook from NC literature survey markdown files."""

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/Users/alok/Desktop/iit hyd/resch/NC_Literature_Survey.xlsx"

NC_CATS = [
    ("NC-1", "Header Field Context"),
    ("NC-2", "Protocol Structure Context"),
    ("NC-3", "Map / Stateful Object Context"),
    ("NC-4", "Packet Action Context"),
    ("NC-5", "Behavioral Contract Context"),
    ("NC-6", "Binary Semantic Context"),
    ("NC-7", "Temporal / Time-Driven Context"),
    ("NC-8", "Resource / Performance Context"),
    ("NC-9", "Chain / Composition Context"),
]

# Coverage: Full, Partial, Taxonomy, None
# YP baseline from full_deep_nc_survey
PAPERS = [
    {
        "id": 1,
        "short_name": "Yaksha-Prashna",
        "full_title": "Yaksha-Prashna: Understanding eBPF Bytecode Network Function Behavior",
        "authors": "Singh et al.",
        "venue": "arXiv",
        "year": 2026,
        "doi": "arXiv:2602.11232",
        "domain": "eBPF / Baseline",
        "target": "eBPF bytecode",
        "method": "Dataflow on eBPF CFG; Prolog KB; rule table",
        "nc": ["Partial", "Partial", "Partial", "Partial", "None", "None", "None", "None", "Partial"],
        "nc_detail": "NC-1a,b; NC-2a,b; NC-3a,c(partial); NC-4a,b; NC-9a",
        "key_features": "readsField, updatesField, accessesProtocol, mapLookup, mapWrite, correlatedMaps, drops, passes, redirects, invokedHelpers",
        "violations": "Prolog: RAW/WAR/WAW; 24 assertion predicates",
        "yp_gain": "Baseline — compare all extensions against this",
        "speed": "13–300 ms per NF",
        "sources": "All 3 surveys",
    },
    {
        "id": 2,
        "short_name": "BinPRE",
        "full_title": "BinPRE: Enhancing Field Inference in Binary Analysis Based Protocol RE",
        "authors": "Jiang et al.",
        "venue": "CCS",
        "year": 2024,
        "doi": "10.1145/3658644.3690299",
        "domain": "Protocol RE",
        "target": "Binary protocol implementations",
        "method": "Instruction semantic similarity; atomic semantic detectors; cluster-refine",
        "nc": ["Full", "Partial", "None", "None", "None", "None", "None", "None", "None"],
        "nc_detail": "NC-1a-d; NC-2b,c",
        "key_features": "Field boundaries; TYPE/LENGTH/SEQ/CHECKSUM/PAYLOAD roles; comparison constants",
        "violations": "PRE accuracy (F1 0.74 type); schema compatibility between NFs",
        "yp_gain": "FieldRole on readsField/updatesField; fixes implicit protocol handling",
        "speed": "N/A",
        "sources": "All 3 surveys",
    },
    {
        "id": 3,
        "short_name": "NetLifter",
        "full_title": "Lifting Network Protocol Implementation to Precise Format Specification",
        "authors": "Shi et al.",
        "venue": "CCS",
        "year": 2023,
        "doi": "10.1145/3576915.3616614",
        "domain": "Protocol RE",
        "target": "Protocol parser source/binary",
        "method": "Abstract interpretation on CFG → Abstract Format Graph (AFG)",
        "nc": ["Full", "Full", "None", "None", "None", "None", "None", "None", "None"],
        "nc_detail": "NC-1a-f; NC-2a-d,f",
        "key_features": "AFG; conditional field presence; cross-field dependencies; 5-tuple key composition",
        "violations": "AFG mismatch between implementations → CVE (SSH)",
        "yp_gain": "Conditional updatesField; mapKey composition; full parse tree",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 4,
        "short_name": "AIFORE",
        "full_title": "AIFORE: Smart Fuzzing Based on Automatic Input Format RE",
        "authors": "Shi et al.",
        "venue": "USENIX Security",
        "year": 2023,
        "doi": "N/A",
        "domain": "Protocol RE / Fuzzing",
        "target": "Network protocol binaries",
        "method": "Byte taint; NN on BB features (opcode histogram, comparisons)",
        "nc": ["Full", "Partial", "None", "None", "None", "Partial", "None", "None", "None"],
        "nc_detail": "NC-1a-d; NC-2a-d; NC-6c",
        "key_features": "Field boundaries 84%; types INTEGER/STRING/ENUM/SIZE/MAGIC/CHECKSUM",
        "violations": "Fuzzer coverage; valid input generation",
        "yp_gain": "BB feature vector for eBPF blocks; protocol discriminant constants",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 5,
        "short_name": "SM Extract (Loop FSM)",
        "full_title": "Extracting Protocol Format as State Machine via Controlled Static Loop Analysis",
        "authors": "Shi, Xu, Zhang",
        "venue": "USENIX Security",
        "year": 2023,
        "doi": "arXiv:2305.13483",
        "domain": "Protocol RE",
        "target": "Parsing loops in protocol implementations",
        "method": "Loop iteration = state; controlled path merging",
        "nc": ["None", "Full", "None", "None", "None", "None", "None", "None", "None"],
        "nc_detail": "NC-2a-e,f (main: NC-2e state machine)",
        "key_features": "Protocol FSM from parsing loops; >90% precision/recall <5 min",
        "violations": "Fuzzer finds 10+ zero-days via extracted SM",
        "yp_gain": "NC-2e for eBPF option-parsing loops — biggest NC-2 gap fix",
        "speed": "<5 min",
        "sources": "full_deep",
    },
    {
        "id": 6,
        "short_name": "Header Space Analysis",
        "full_title": "Header Space Analysis: Static Checking for Networks",
        "authors": "Kazemian, Varghese, McKeown",
        "venue": "NSDI",
        "year": 2012,
        "doi": "N/A",
        "domain": "Network verification",
        "target": "Forwarding tables / ACLs (manual)",
        "method": "Header space {0,1}^L; box transfer functions; composition",
        "nc": ["Partial", "None", "None", "Partial", "None", "None", "None", "None", "Partial"],
        "nc_detail": "NC-1a,d; NC-4a,b; NC-9a",
        "key_features": "Bit-level header space; reachability; isolation; forwarding loops",
        "violations": "Header space intersection; slice leak; loop fixed point",
        "yp_gain": "Transfer function composition for chain analysis without path explosion",
        "speed": "N/A",
        "sources": "All 3 surveys",
    },
    {
        "id": 7,
        "short_name": "NetPlumber",
        "full_title": "Real Time Network Policy Checking Using Header Space Analysis",
        "authors": "Kazemian et al.",
        "venue": "NSDI",
        "year": 2013,
        "doi": "N/A",
        "domain": "Network verification",
        "target": "Dynamic network policies",
        "method": "Incremental HSA updates; rule dependency graph",
        "nc": ["Partial", "None", "None", "Partial", "None", "None", "None", "None", "Partial"],
        "nc_detail": "NC-1a-d; NC-4a,b; NC-9a",
        "key_features": "Incremental transfer functions; rule conflicts; dependency loops",
        "violations": "Rule conflict; dependency loop in header space",
        "yp_gain": "Incremental chain recompute when NF added to pipeline",
        "speed": "Real-time",
        "sources": "categorised + full_deep",
    },
    {
        "id": 8,
        "short_name": "PREVAIL",
        "full_title": "Simple and Precise Static Analysis of Untrusted Linux Kernel Extensions",
        "authors": "Gershuni et al.",
        "venue": "PLDI",
        "year": 2019,
        "doi": "10.1145/3314221.3314590",
        "domain": "eBPF safety",
        "target": "eBPF programs",
        "method": "Abstract interpretation: Offset_value + Interval domains (Crab)",
        "nc": ["Partial", "Partial", "Partial", "Partial", "None", "None", "None", "None", "None"],
        "nc_detail": "NC-1a,b; NC-2a-c; NC-3b,c; NC-4a",
        "key_features": "PTR_TO_PACKET/MAP_VALUE/STACK/CTX; interval ranges; loop bounds",
        "violations": "OOB access; illegal pointer arithmetic (safety only)",
        "yp_gain": "Memory region typing — fixes NF12 UDP implicit protocol bug",
        "speed": "Fast (kernel verifier scale)",
        "sources": "All 3 surveys",
    },
    {
        "id": 9,
        "short_name": "eBPF Bit-Precise Verify",
        "full_title": "Automatic Bit-and Memory-Precise Verification of eBPF Code",
        "authors": "Bromberger, Schwarz, Weidenbach",
        "venue": "LPAR",
        "year": 2024,
        "doi": "N/A",
        "domain": "eBPF formal verification",
        "target": "eBPF bytecode",
        "method": "Translation to FOL + bitvectors; Isabelle/HOL",
        "nc": ["None", "None", "Partial", "None", "None", "None", "None", "None", "None"],
        "nc_detail": "NC-3 (memory typing); formal ISA",
        "key_features": "Bitvector-precise eBPF ISA; soundness proof",
        "violations": "Memory safety; functional correctness vs spec",
        "yp_gain": "Formal soundness for YP dataflow rules",
        "speed": "N/A",
        "sources": "network_context",
    },
    {
        "id": 10,
        "short_name": "BinaryInferno",
        "full_title": "BinaryInferno: A Semantic-Driven Approach to Field Inference",
        "authors": "Chandler",
        "venue": "IEEE CNS / WOOT",
        "year": 2023,
        "doi": "N/A",
        "domain": "Protocol RE",
        "target": "Binary message formats",
        "method": "Delimiter detection; semantic-driven inference",
        "nc": ["Partial", "Partial", "None", "None", "None", "None", "None", "None", "None"],
        "nc_detail": "NC-1a,c; NC-2b",
        "key_features": "Delimiter bytes = branch comparison constants in eBPF",
        "violations": "PRE field boundary accuracy",
        "yp_gain": "Extract full protocol discriminant set from branches",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 11,
        "short_name": "Protocol RE Survey",
        "full_title": "Protocol Reverse-Engineering Methods and Tools: A Survey",
        "authors": "Kleber, Kopp, Kargl",
        "venue": "Computer Communications",
        "year": 2021,
        "doi": "10.1016/j.comcom.2021.11.009",
        "domain": "Survey / Taxonomy",
        "target": "PRE literature",
        "method": "6-phase PRE taxonomy",
        "nc": ["Taxonomy", "Taxonomy", "None", "None", "Taxonomy", "None", "None", "None", "None"],
        "nc_detail": "Maps PRE phases → NC categories",
        "key_features": "Phases: segmentation, type ID, semantics, state machine, behavior model",
        "violations": "N/A (survey)",
        "yp_gain": "Phase 4–5 = unsolved for eBPF (state machine + behavior model)",
        "speed": "N/A",
        "sources": "All 3 surveys",
    },
    {
        "id": 12,
        "short_name": "Klint",
        "full_title": "Automated Verification of Network Function Binaries",
        "authors": "Pirelli et al.",
        "venue": "NSDI",
        "year": 2022,
        "doi": "10.usenix.org/conference/nsdi22/presentation/pirelli",
        "domain": "NF binary verification",
        "target": "NF binaries (no source)",
        "method": "KLEE symbolic execution + VeriFast; ghost map abstraction",
        "nc": ["None", "None", "Full", "Partial", "Full", "None", "None", "None", "Partial"],
        "nc_detail": "NC-3a,c-g; NC-4a,b,f; NC-5a-d; NC-9b,c",
        "key_features": "lookup/insert/remove/size/contains; key/value schema; null→drop/pass/insert",
        "violations": "Python spec vs ghost map op sequence mismatch",
        "yp_gain": "mapOp(NF,M,LOOKUP/INSERT/REMOVE,key_fields,null_behavior)",
        "speed": "2.7s–1.5 min per NF (200–1000× slower than YP)",
        "sources": "All 3 surveys",
    },
    {
        "id": 13,
        "short_name": "Vigor",
        "full_title": "Verifying Software Network Functions with No Verification Expertise",
        "authors": "Zaostrovnykh et al.",
        "venue": "SOSP",
        "year": 2019,
        "doi": "N/A",
        "domain": "NF verification",
        "target": "Software NFs (LibVig)",
        "method": "KLEE paths + VeriFast stateful libs; behavioral summary",
        "nc": ["None", "None", "Full", "Partial", "Full", "None", "None", "None", "Partial"],
        "nc_detail": "NC-3a,c-g; NC-4a,b,f; NC-5a,b,d,e; NC-9b",
        "key_features": "Behavioral summary = map op sequence per path; LibVig contracts",
        "violations": "Path where stateful call sequence ≠ Python spec",
        "yp_gain": "behaviorPattern(NF, [lookup→if null drop else forward])",
        "speed": "Minutes per NF",
        "sources": "All 3 surveys",
    },
    {
        "id": 14,
        "short_name": "NetSMC",
        "full_title": "NetSMC: Stateful Network Verification via Symbolic Model Checking",
        "authors": "Yuan et al.",
        "venue": "NSDI",
        "year": 2020,
        "doi": "usenix.org/conference/nsdi20/presentation/yuan",
        "domain": "Stateful network verification",
        "target": "Policy language (not bytecode)",
        "method": "Backward reachability; custom SMC O(states×NFs)",
        "nc": ["Partial", "None", "Full", "Partial", "Partial", "None", "Partial", "None", "Full"],
        "nc_detail": "NC-3a,c,d,e,g; NC-4a,b,f; NC-5b,d; NC-7c,d; NC-9b,c,d",
        "key_features": "state_table[flow_key]=state; transition rules; location model",
        "violations": "Bad state reachable → witness trace",
        "yp_gain": "Connect mapWrite to state transition model",
        "speed": "Polynomial relaxations",
        "sources": "All 3 surveys",
    },
    {
        "id": 15,
        "short_name": "Alpernas SAS 2018",
        "full_title": "Abstract Interpretation of Stateful Networks",
        "authors": "Alpernas et al.",
        "venue": "SAS",
        "year": 2018,
        "doi": "10.1007/978-3-319-99725-4_3",
        "domain": "Stateful network verification",
        "target": "Middlebox specifications",
        "method": "Packet effect semantics; Cartesian abstraction",
        "nc": ["None", "None", "Partial", "Full", "Partial", "None", "None", "None", "Full"],
        "nc_detail": "NC-3c,e,g,h; NC-4d,e,f; NC-5b,d,e; NC-9b,c,d,f",
        "key_features": "Per-packet-type effect; O(n^k) k≤5; Class 1/2/3 complexity",
        "violations": "Isolation via abstract interpretation fixpoint",
        "yp_gain": "Packet-effect automaton; polynomial chain verification",
        "speed": "Polynomial for real NFs",
        "sources": "All 3 surveys",
    },
    {
        "id": 16,
        "short_name": "Alpernas FMCS 2021",
        "full_title": "Some Complexity Results for Stateful Network Verification",
        "authors": "Alpernas et al.",
        "venue": "FMCS",
        "year": 2021,
        "doi": "arXiv:2106.01030",
        "domain": "Theory",
        "target": "Middlebox classification",
        "method": "Complexity proofs",
        "nc": ["None", "None", "Partial", "None", "None", "None", "None", "None", "None"],
        "nc_detail": "NC-3h only",
        "key_features": "Class 1 POLY / Class 2 coNP / Class 3 EXPSPACE / General UNDECIDABLE",
        "violations": "Tractability classification before querying",
        "yp_gain": "Infer NC class from null_behavior + key schema",
        "speed": "N/A",
        "sources": "All 3 surveys",
    },
    {
        "id": 17,
        "short_name": "eBPF Maps Workshop",
        "full_title": "Understanding Performance of eBPF Maps",
        "authors": "SIGCOMM eBPF Workshop",
        "venue": "SIGCOMM Workshop",
        "year": 2024,
        "doi": "N/A",
        "domain": "eBPF maps",
        "target": "eBPF map types",
        "method": "Performance taxonomy from ELF descriptors",
        "nc": ["None", "None", "Partial", "None", "None", "None", "Partial", "Partial", "None"],
        "nc_detail": "NC-3b; NC-7e; NC-8b,c",
        "key_features": "HASH/ARRAY/LRU/LPM/PERCPU/RINGBUF/SOCKHASH semantics",
        "violations": "N/A",
        "yp_gain": "Map type from ELF at zero analysis cost",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 18,
        "short_name": "Aquila",
        "full_title": "Aquila: Verification for Production-Scale Programmable Data Planes",
        "authors": "Tian et al.",
        "venue": "SIGCOMM",
        "year": 2021,
        "doi": "10.1145/3452296.3472889",
        "domain": "P4 verification",
        "target": "P4 programs",
        "method": "Sequential SMT encoding; existential over table entries",
        "nc": ["Partial", "Partial", "None", "Full", "None", "None", "None", "None", "Partial"],
        "nc_detail": "NC-1a,b; NC-2a-d; NC-4a-d,f; NC-9a,c,d",
        "key_features": "Table lookup as ∃ entry; undefined behavior detection; bug localization",
        "violations": "SAT/UNSAT ∃ packet violating assertion",
        "yp_gain": "Sequential encoding for eBPF map lookups (avoid path explosion)",
        "speed": "Production-scale P4",
        "sources": "All 3 surveys",
    },
    {
        "id": 19,
        "short_name": "PIX",
        "full_title": "Performance Interfaces for Network Functions",
        "authors": "Iyer et al.",
        "venue": "NSDI",
        "year": 2022,
        "doi": "N/A",
        "domain": "NF performance",
        "target": "eBPF XDP NFs (Katran, Cilium, Natasha)",
        "method": "Symbolic analysis; path coverage per packet type",
        "nc": ["Partial", "None", "Partial", "Full", "Partial", "None", "None", "Full", "None"],
        "nc_detail": "NC-1a,b; NC-3d; NC-4b,d; NC-5e; NC-8a-d",
        "key_features": "Path per packet type; instruction count; cache footprint",
        "violations": "Performance contract violation per packet type",
        "yp_gain": "Concrete packet-effect paths from eBPF bytecode",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 20,
        "short_name": "BOLT",
        "full_title": "Performance Contracts for Software Network Functions",
        "authors": "Iyer et al.",
        "venue": "NSDI",
        "year": 2019,
        "doi": "N/A",
        "domain": "NF performance",
        "target": "Software NFs + DPDK stack",
        "method": "Critical variable analysis; full stack modeling",
        "nc": ["None", "None", "Partial", "Partial", "Partial", "None", "None", "Full", "None"],
        "nc_detail": "NC-3d; NC-4d; NC-5e; NC-8a-e",
        "key_features": "num_flows, num_backends as critical vars; perf = f(critical vars)",
        "violations": "Latency exceeds contract when critical var threshold crossed",
        "yp_gain": "Critical variables = map capacity → behavior change on full map",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 21,
        "short_name": "Jones Rely-Guarantee",
        "full_title": "Tentative Steps Toward a Development Method for Interfering Programs",
        "authors": "Jones, C.B.",
        "venue": "ACM TOPLAS",
        "year": 1983,
        "doi": "N/A",
        "domain": "Formal methods",
        "target": "Concurrent programs",
        "method": "Rely/Guarantee framework",
        "nc": ["None", "None", "None", "None", "Full", "None", "None", "None", "None"],
        "nc_detail": "NC-5b-d",
        "key_features": "Rely(P), Guarantee(P); chain: Guarantee(i) ⊇ Rely(i+1)",
        "violations": "Rely/Guarantee mismatch (Cilium+AWS outage formalized)",
        "yp_gain": "Formal NC-5 foundation for NF chains",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 22,
        "short_name": "Feng Local RG",
        "full_title": "On the Relationship Between Concurrent Separation Logic and Assume-Guarantee",
        "authors": "Feng et al.",
        "venue": "POPL",
        "year": 2007,
        "doi": "N/A",
        "domain": "Formal methods",
        "target": "Programs with shared state",
        "method": "Backward analysis from branches → implicit Rely",
        "nc": ["None", "None", "None", "None", "Full", "None", "None", "None", "None"],
        "nc_detail": "NC-5b-d",
        "key_features": "Automated Rely: branch on field F → Rely(F valid)",
        "violations": "Guarantee ⊉ Rely at chain link",
        "yp_gain": "Automated Rely extraction via YP backward pass",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 23,
        "short_name": "VEP",
        "full_title": "VEP: Two-Stage Verification Toolchain for Full eBPF Programmability",
        "authors": "Wu et al.",
        "venue": "NSDI",
        "year": 2025,
        "doi": "N/A",
        "domain": "eBPF verification",
        "target": "eBPF programs (C annotated)",
        "method": "Proof-carrying code; cross-program invariants",
        "nc": ["None", "None", "Partial", "None", "Full", "None", "None", "None", "None"],
        "nc_detail": "NC-3; NC-5b-d",
        "key_features": "Cross-program invariants for eBPF chains",
        "violations": "Formal proof failure on chain invariant",
        "yp_gain": "NF_i postcondition = NF_{i+1} precondition model",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 24,
        "short_name": "CLAP",
        "full_title": "CLAP: Contrastive Binary-NL Alignment (project)",
        "authors": "Project file",
        "venue": "Project",
        "year": 0,
        "doi": "N/A",
        "domain": "Binary understanding",
        "target": "Stripped binaries",
        "method": "Contrastive learning; BB embeddings; lifted IR",
        "nc": ["None", "None", "None", "None", "Partial", "Full", "None", "None", "None"],
        "nc_detail": "NC-5e; NC-6b,c,e",
        "key_features": "64-dim BB embedding; 83% Recall@1 zero-shot NF role",
        "violations": "N/A",
        "yp_gain": "Semantic labels: flow affinity, stateful firewall",
        "speed": "N/A",
        "sources": "All 3 surveys",
    },
    {
        "id": 25,
        "short_name": "Bin2Summary",
        "full_title": "Bin2Summary: Argument Data Flow Slicing (project)",
        "authors": "Project file",
        "venue": "Project",
        "year": 0,
        "doi": "N/A",
        "domain": "Binary understanding",
        "target": "Binary functions",
        "method": "Forward slice from input argument (r1=xdp_md)",
        "nc": ["None", "None", "None", "None", "Partial", "Full", "None", "None", "None"],
        "nc_detail": "NC-5a; NC-6b,d",
        "key_features": "Packet blocks vs bookkeeping blocks separation",
        "violations": "N/A",
        "yp_gain": "Slice r1→NC-1/2/4; slice r6→NC-3",
        "speed": "N/A",
        "sources": "All 3 surveys",
    },
    {
        "id": 26,
        "short_name": "Neural RE",
        "full_title": "Neural Reverse Engineering of Stripped Binaries",
        "authors": "David, Alon, Yahav",
        "venue": "OOPSLA",
        "year": 2020,
        "doi": "arXiv:1902.09122",
        "domain": "Binary understanding",
        "target": "Stripped binaries",
        "method": "API call sequences + argument types → name prediction",
        "nc": ["None", "None", "None", "None", "None", "Full", "None", "None", "None"],
        "nc_detail": "NC-6a",
        "key_features": "BPF helper sequences predict NF role (LB, NAT, etc.)",
        "violations": "N/A",
        "yp_gain": "Enrich invokedHelpers with argument types",
        "speed": "81.7% precision",
        "sources": "categorised + full_deep",
    },
    {
        "id": 27,
        "short_name": "Gemini",
        "full_title": "Neural Graph Embedding for Cross-Platform Binary Code Similarity",
        "authors": "Xu et al.",
        "venue": "CCS",
        "year": 2017,
        "doi": "N/A",
        "domain": "Binary similarity",
        "target": "Cross-platform binaries",
        "method": "ACFG + GNN → 64-dim embedding",
        "nc": ["None", "None", "None", "None", "None", "Full", "None", "None", "None"],
        "nc_detail": "NC-6c",
        "key_features": "ACFG: op counts, comparison constants, memory accesses",
        "violations": "N/A",
        "yp_gain": "NF clustering / bisimulation fingerprint",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 28,
        "short_name": "jTrans",
        "full_title": "jTrans: Jump-Aware Transformer for Binary Code Similarity",
        "authors": "Wang et al.",
        "venue": "ISSTA",
        "year": 2022,
        "doi": "arXiv:2205.12713",
        "domain": "Binary similarity",
        "target": "Assembly binaries",
        "method": "Jump-aware BERT on instruction sequences",
        "nc": ["None", "None", "None", "None", "None", "Full", "None", "None", "None"],
        "nc_detail": "NC-6c",
        "key_features": "Long-range branch context for map lookup usage",
        "violations": "N/A",
        "yp_gain": "Context before conditional branches in eBPF",
        "speed": "N/A",
        "sources": "categorised + full_deep",
    },
    {
        "id": 29,
        "short_name": "Gigahorse",
        "full_title": "Gigahorse: Declarative Decompilation of Smart Contracts",
        "authors": "Grech et al.",
        "venue": "ICSE",
        "year": 2019,
        "doi": "10.1145/3338906.3338977",
        "domain": "Bytecode analysis",
        "target": "EVM bytecode",
        "method": "Soufflé Datalog decompilation rules",
        "nc": ["None", "None", "Partial", "None", "None", "Full", "None", "None", "None"],
        "nc_detail": "NC-3a; NC-6 all",
        "key_features": "Datalog IR; 10–100× faster than Prolog; provenance",
        "violations": "N/A",
        "yp_gain": "Migrate YP SWI-Prolog → Soufflé Datalog",
        "speed": "10–100× vs Prolog",
        "sources": "categorised + full_deep",
    },
    {
        "id": 30,
        "short_name": "NF-SE",
        "full_title": "Symbolic Execution for Network Functions with Time-Driven Logic",
        "authors": "MASCOTS authors",
        "venue": "IEEE MASCOTS",
        "year": 2020,
        "doi": "10.1109/MASCOTS50786.2020.9285941",
        "domain": "Temporal NF verification",
        "target": "SEFL network functions",
        "method": "Symbolic execution + time primitives",
        "nc": ["None", "None", "None", "Partial", "None", "None", "Full", "None", "None"],
        "nc_detail": "NC-4d; NC-7a-d",
        "key_features": "on_timeout, timer_reset; multi-packet sequences; ts field",
        "violations": "Temporal predicate violations (response within 150ms)",
        "yp_gain": "Patterns after bpf_ktime_get_ns in SSA",
        "speed": "N/A",
        "sources": "full_deep",
    },
    {
        "id": 31,
        "short_name": "Hydra",
        "full_title": "Hydra: Effective Runtime Network Verification",
        "authors": "Renganathan et al.",
        "venue": "SIGCOMM",
        "year": 2023,
        "doi": "10.1145/3603269.3604856",
        "domain": "Runtime verification",
        "target": "P4 data planes",
        "method": "Indus DSL → P4 line-rate monitors; LTL_f",
        "nc": ["None", "None", "None", "Full", "None", "None", "Full", "None", "Full"],
        "nc_detail": "NC-4a-e; NC-7c,d; NC-9c,d,f",
        "key_features": "Chain ordering; multi-packet deps; witness on violation",
        "violations": "Runtime LTL_f property violation with packet capture",
        "yp_gain": "NC-9c ordering constraints; NC-7d multi-packet",
        "speed": "Line rate",
        "sources": "full_deep",
    },
    {
        "id": 32,
        "short_name": "MFOTL Runtime",
        "full_title": "Compiling Stateful Network Properties for Runtime Verification",
        "authors": "arXiv authors",
        "venue": "arXiv",
        "year": 2016,
        "doi": "arXiv:1607.03385",
        "domain": "Temporal verification",
        "target": "Network switches",
        "method": "MFOTL → switch-executable rules",
        "nc": ["None", "None", "None", "None", "None", "None", "Full", "None", "Partial"],
        "nc_detail": "NC-7c-e; NC-9c,e",
        "key_features": "Time-indexed predicates; state expiry at line rate",
        "violations": "Event reordering; timeout violations",
        "yp_gain": "LRU map + null behavior = NC-7e temporal dep",
        "speed": "Line rate",
        "sources": "full_deep",
    },
    {
        "id": 33,
        "short_name": "Temporal Logic Patent",
        "full_title": "Verify a Network Function via Query Language with Temporal Logic",
        "authors": "US Patent",
        "venue": "US Patent",
        "year": 0,
        "doi": "US10958547",
        "domain": "NF temporal verification",
        "target": "Mixed source/binary NFs",
        "method": "LTL past/future operators on NF state machine",
        "nc": ["None", "None", "None", "None", "None", "None", "Full", "None", "Partial"],
        "nc_detail": "NC-7a-d; NC-9c,d",
        "key_features": "EF, AG, EP, AP temporal operators",
        "violations": "Temporal property violations in NF chain",
        "yp_gain": "Query language for temporal NC beyond static YP",
        "speed": "N/A",
        "sources": "full_deep",
    },
    {
        "id": 34,
        "short_name": "SymNet",
        "full_title": "SymNet: Scalable Symbolic Execution for Modern Networks",
        "authors": "Stoenescu et al.",
        "venue": "SIGCOMM",
        "year": 2016,
        "doi": "10.1145/2934872.2934881",
        "domain": "Network symbolic execution",
        "target": "Click routers / SEFL",
        "method": "SEFL language; exact branching",
        "nc": ["Partial", "Full", "Partial", "Full", "None", "None", "None", "None", "Full"],
        "nc_detail": "NC-1a,b; NC-2a-d; NC-3 partial; NC-4a-d; NC-9a,c,d,f",
        "key_features": "Stateful processing; tunneling; no false paths",
        "violations": "Forwarding loops, black holes, NAT bugs",
        "yp_gain": "Exact branching model for eBPF (no false positives)",
        "speed": "Scalable symbolic exec",
        "sources": "full_deep",
    },
    {
        "id": 35,
        "short_name": "NetKAT",
        "full_title": "NetKAT: Semantic Foundations for Networks",
        "authors": "Anderson et al.",
        "venue": "POPL",
        "year": 2014,
        "doi": "N/A",
        "domain": "Network algebra",
        "target": "Network programs",
        "method": "Kleene algebra with tests; equational reasoning",
        "nc": ["Partial", "None", "None", "Partial", "None", "None", "Partial", "None", "Full"],
        "nc_detail": "NC-1a,b; NC-4a,b,e; NC-7c; NC-9a,c,d,e",
        "key_features": "filter, modify, union, sequence, iteration primitives",
        "violations": "Reachability, isolation, non-interference via algebra",
        "yp_gain": "Formal sound composition operator for YP chains",
        "speed": "N/A",
        "sources": "full_deep",
    },
    {
        "id": 36,
        "short_name": "VeriFlow",
        "full_title": "VeriFlow: Verifying Network-Wide Invariants in Real Time",
        "authors": "Khurshid et al.",
        "venue": "NSDI",
        "year": 2013,
        "doi": "N/A",
        "domain": "Real-time verification",
        "target": "Network forwarding rules",
        "method": "Equivalence classes; incremental checking",
        "nc": ["Partial", "None", "None", "Full", "None", "None", "None", "None", "Full"],
        "nc_detail": "NC-1a,b; NC-4a,b,c,e; NC-9a,c,d",
        "key_features": "Per-EC violation check; μs per rule update",
        "violations": "Reachability, loops, isolation per EC",
        "yp_gain": "Incremental chain invariant on NF insertion",
        "speed": "Hundreds of μs",
        "sources": "full_deep",
    },
    {
        "id": 37,
        "short_name": "Middlebox Patent",
        "full_title": "Middlebox Modeling (Packet Processing Slice)",
        "authors": "US Patent",
        "venue": "US Patent",
        "year": 0,
        "doi": "US10594574",
        "domain": "Middlebox analysis",
        "target": "Source code middleboxes",
        "method": "Packet/state/config variable categorization; backward slice",
        "nc": ["None", "None", "Partial", "Partial", "Full", "None", "None", "None", "Full"],
        "nc_detail": "NC-3b; NC-4a,b; NC-5a-d; NC-9a-d",
        "key_features": "3 variable categories; slice from forwarding action",
        "violations": "Chain state inconsistency",
        "yp_gain": "Separate r1 (packet) vs r6 (map) registers in eBPF",
        "speed": "N/A",
        "sources": "full_deep",
    },
    {
        "id": 38,
        "short_name": "CoNEXT eBPF Perf",
        "full_title": "Demystifying Performance of eBPF Network Applications",
        "authors": "NYU et al.",
        "venue": "CoNEXT",
        "year": 2025,
        "doi": "N/A",
        "domain": "eBPF performance",
        "target": "eBPF network apps",
        "method": "Map access patterns; JIT behavior; memory footprint",
        "nc": ["None", "None", "Partial", "None", "None", "None", "Partial", "Full", "None"],
        "nc_detail": "NC-3b,d; NC-7e; NC-8b,c,d",
        "key_features": "Map access time; read-heavy vs write-heavy maps",
        "violations": "N/A",
        "yp_gain": "Distinguish routing table vs conntrack by access pattern",
        "speed": "N/A",
        "sources": "full_deep",
    },
    {
        "id": 39,
        "short_name": "Panda Middleboxes",
        "full_title": "Verifying Isolation Properties in the Presence of Middleboxes",
        "authors": "Panda et al.",
        "venue": "NSDI",
        "year": 2015,
        "doi": "N/A",
        "domain": "Middlebox verification",
        "target": "Middlebox-enhanced networks",
        "method": "Header space + middlebox models",
        "nc": ["Partial", "None", "Partial", "Partial", "None", "None", "None", "None", "Partial"],
        "nc_detail": "NC-1; NC-3; NC-4; NC-9",
        "key_features": "Isolation with middleboxes in path",
        "violations": "Isolation leakage through middlebox",
        "yp_gain": "Early middlebox chain verification foundation",
        "speed": "N/A",
        "sources": "full_deep ref list",
    },
    {
        "id": 40,
        "short_name": "Soufflé",
        "full_title": "Soufflé: On Synthesis of Program Analyzers in Datalog",
        "authors": "Jordan et al.",
        "venue": "CAV",
        "year": 2016,
        "doi": "N/A",
        "domain": "Analysis infrastructure",
        "target": "Datalog analyzers",
        "method": "Datalog → optimized C++",
        "nc": ["None", "None", "None", "None", "None", "None", "None", "None", "None"],
        "nc_detail": "Infrastructure for Gigahorse-style YP migration",
        "key_features": "Fast Datalog engine; provenance tracking",
        "violations": "N/A",
        "yp_gain": "Backend for 10–100× faster YP queries",
        "speed": "10–100× vs Prolog",
        "sources": "full_deep ref list",
    },
]

YP_GAPS = [
    ("NC-1", "Header Field", "NC-1c field role", "BinPRE atomic detectors", "High"),
    ("NC-1", "Header Field", "NC-1d value ranges", "BinPRE, NetLifter", "Medium"),
    ("NC-1", "Header Field", "NC-1e conditional writes", "NetLifter AFG", "High"),
    ("NC-1", "Header Field", "NC-1f field composition (5-tuple keys)", "NetLifter AFG", "High"),
    ("NC-2", "Protocol Structure", "NC-2c implicit protocol (NF12 UDP bug)", "PREVAIL memory regions", "Critical"),
    ("NC-2", "Protocol Structure", "NC-2d full parse tree", "NetLifter, Aquila", "High"),
    ("NC-2", "Protocol Structure", "NC-2e parsing loop state machine", "USENIX Sec'23 loop analysis", "High"),
    ("NC-2", "Protocol Structure", "NC-2f unhandled packet types", "Aquila sequential encoding", "Medium"),
    ("NC-3", "Map / State", "NC-3b map type from ELF", "SIGCOMM eBPF Maps Workshop", "Low (zero cost)"),
    ("NC-3", "Map / State", "NC-3d key schema", "Klint SSA backward slice", "Critical"),
    ("NC-3", "Map / State", "NC-3e value type / semantics", "Klint ghost map", "Critical"),
    ("NC-3", "Map / State", "NC-3f null behavior", "Klint branch patterns", "Critical"),
    ("NC-3", "Map / State", "NC-3g cross-NF map consistency", "Klint, NetSMC", "High"),
    ("NC-3", "Map / State", "NC-3h complexity class", "Alpernas FMCS 2021", "Medium"),
    ("NC-4", "Packet Action", "NC-4c action completeness / gaps", "Aquila ∃ encoding", "High"),
    ("NC-4", "Packet Action", "NC-4d per-packet-type action", "Alpernas packet effect, PIX", "High"),
    ("NC-4", "Packet Action", "NC-4e chain action composition", "NetKAT sequence operator", "High"),
    ("NC-4", "Packet Action", "NC-4f action+state mutation pairs", "Vigor behavioral summary", "Medium"),
    ("NC-5", "Behavioral Contract", "Entire NC-5 absent", "Jones RG + Feng + Klint spec", "Critical"),
    ("NC-6", "Binary Semantic", "Entire NC-6 absent", "CLAP, Bin2Summary, Neural RE", "Medium"),
    ("NC-7", "Temporal", "Entire NC-7 absent", "NF-SE, Hydra, bpf_ktime patterns", "High"),
    ("NC-8", "Resource", "Entire NC-8 absent", "PIX, BOLT on eBPF XDP", "Low-Medium"),
    ("NC-9", "Chain", "NC-9b map-level deps", "Klint, NetSMC", "Critical"),
    ("NC-9", "Chain", "NC-9c ordering constraints", "Hydra, NetSMC policy", "High"),
    ("NC-9", "Chain", "NC-9d action completeness in chain", "NetKAT, VeriFlow", "High"),
    ("NC-9", "Chain", "NC-9e resource contention", "BOLT", "Low"),
    ("NC-9", "Chain", "NC-9f emergent chain behavior", "Alpernas, SymNet", "High"),
]

SYNTHESIS_GAPS = [
    ("Map operation type", "lookup-then-drop vs lookup-then-redirect", "Trace map_lookup result to branch", "Klint ghost map ops"),
    ("Map key schema", "Which fields form map key", "Backward SSA from map helper", "BinPRE field roles"),
    ("Map value structure", "What stored value means", "Forward trace of lookup result", "Klint ghost map"),
    ("Field semantic role", "TYPE/LENGTH/SEQ/CHECKSUM", "BinPRE atomic detectors", "BinPRE CCS'24"),
    ("NF state machine", "States and transitions", "Map key→value→branch pattern", "NetSMC"),
    ("Behavioral pattern", "Firewall/LB/router contract", "Map ops + packet actions", "Vigor"),
    ("Packet effect semantics", "Per packet type effect", "CFG path per packet type", "Alpernas SAS'18"),
    ("Protocol skip behavior", "Implicit protocol handling", "Field role not equality check", "PREVAIL + BinPRE"),
    ("Chain state evolution", "Shared map across NFs", "Packet-effect composition", "Alpernas + NetSMC"),
    ("Runtime assumptions", "Environment preconditions", "Pattern match on helpers", "SemanticIR design"),
]

PROBLEM_MAPPING = [
    ("Problem 1: Path explosion", "Lack of NF summary features composable without re-enumerating paths", "Transfer function + map op type", "Klint, PREVAIL"),
    ("Problem 2: Stateful queries", "Lack of map semantic features (key, value, op, transition)", "mapOp LOOKUP/INSERT/REMOVE + null_behavior", "Klint, NetSMC, Alpernas"),
    ("Problem 3: Complex chain queries", "Lack of behavioral contract features", "Packet-effect-per-type + behavioral pattern", "Vigor, Alpernas, BinPRE"),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN = Side(style="thin")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, max_w=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_w)


def cov_score(val):
    return {"Full": 3, "Partial": 2, "Taxonomy": 1, "None": 0}.get(val, 0)


def main():
    wb = Workbook()

    # --- Sheet 1: Overview ---
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov["A1"] = "Network Context Literature Survey — Excel Summary"
    ws_ov["A1"].font = Font(bold=True, size=14)
    ws_ov["A3"] = "Source documents:"
    for i, src in enumerate(
        [
            "network_context_extraction_survey.md (14 papers, 6 domains)",
            "categorised_nc_literature_survey.md (27 papers, 6 NC categories)",
            "full_deep_nc_survey.md (40+ papers, 9 NC categories)",
        ],
        4,
    ):
        ws_ov[f"A{i}"] = f"  • {src}"
    ws_ov["A8"] = "Research question:"
    ws_ov["A9"] = "How to extract MORE network context from bytecode than Yaksha-Prashna?"
    ws_ov["A11"] = "9 Network Context Categories:"
    ws_ov.append([])
    ws_ov.append(["Code", "Category", "YP Coverage (full_deep survey)"])
    style_header(ws_ov, ws_ov.max_row)
    yp_cov = [
        ("NC-1", "Header Field Context", "Partial — field names, no roles"),
        ("NC-2", "Protocol Structure Context", "Partial — strict check assumption"),
        ("NC-3", "Map / Stateful Object Context", "Partial — map name only"),
        ("NC-4", "Packet Action Context", "Yes — drops/passes/redirects"),
        ("NC-5", "Behavioral Contract Context", "Absent"),
        ("NC-6", "Binary Semantic Context", "Absent"),
        ("NC-7", "Temporal / Time-Driven Context", "Absent"),
        ("NC-8", "Resource / Performance Context", "Absent"),
        ("NC-9", "Chain / Composition Context", "Partial — field RAW/WAR/WAW only"),
    ]
    for row in yp_cov:
        ws_ov.append(row)
    auto_width(ws_ov)

    # --- Sheet 2: Papers Master ---
    ws_p = wb.create_sheet("Papers_Master")
    p_headers = [
        "ID", "Short Name", "Full Title", "Authors", "Venue", "Year", "DOI/arXiv",
        "Domain", "Analysis Target", "Extraction Method", "NC Coverage Detail",
        "Key Features Extracted", "Violation Detection", "What YP Gains", "Speed/Scale", "Source Surveys",
    ]
    ws_p.append(p_headers)
    style_header(ws_p)
    for p in PAPERS:
        ws_p.append([
            p["id"], p["short_name"], p["full_title"], p["authors"], p["venue"], p["year"],
            p["doi"], p["domain"], p["target"], p["method"], p["nc_detail"],
            p["key_features"], p["violations"], p["yp_gain"], p["speed"], p["sources"],
        ])
    for r in range(2, ws_p.max_row + 1):
        if r % 2 == 0:
            for c in range(1, len(p_headers) + 1):
                ws_p.cell(r, c).fill = ALT_FILL
    auto_width(ws_p, 45)
    ws_p.freeze_panes = "A2"

    # --- Sheet 3: NC Coverage Matrix ---
    ws_m = wb.create_sheet("NC_Coverage_Matrix")
    m_headers = ["ID", "Paper"] + [c[0] for c in NC_CATS] + ["Coverage Score", "Categories Covered"]
    ws_m.append(m_headers)
    style_header(ws_m)
    for p in PAPERS:
        score = sum(cov_score(v) for v in p["nc"])
        covered = sum(1 for v in p["nc"] if v in ("Full", "Partial", "Taxonomy"))
        ws_m.append([p["id"], p["short_name"]] + p["nc"] + [score, covered])
    # YP row highlight
    yp_row = 2
    for c in range(1, len(m_headers) + 1):
        ws_m.cell(yp_row, c).font = Font(bold=True)
    auto_width(ws_m)
    ws_m.freeze_panes = "C2"

    # Conditional formatting via fill for Full/Partial
    fills = {
        "Full": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Partial": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Taxonomy": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "None": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }
    for r in range(2, ws_m.max_row + 1):
        for c in range(3, 12):
            v = ws_m.cell(r, c).value
            if v in fills:
                ws_m.cell(r, c).fill = fills[v]

    # Chart: papers per NC category (count Full+Partial)
    ws_chart_data = wb.create_sheet("_ChartData")
    ws_chart_data.sheet_state = "hidden"
    ws_chart_data.append(["NC Category", "Papers with Full", "Papers with Partial+Full"])
    for i, (code, name) in enumerate(NC_CATS):
        full = sum(1 for p in PAPERS if p["nc"][i] == "Full")
        any_cov = sum(1 for p in PAPERS if p["nc"][i] in ("Full", "Partial", "Taxonomy"))
        ws_chart_data.append([code, full, any_cov])

    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Papers Covering Each NC Category"
    chart1.y_axis.title = "Number of Papers"
    chart1.x_axis.title = "NC Category"
    data = Reference(ws_chart_data, min_col=3, min_row=1, max_row=10)
    cats = Reference(ws_chart_data, min_col=1, min_row=2, max_row=10)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.width = 18
    chart1.height = 10
    ws_m.add_chart(chart1, "N2")

    # Chart: top papers by coverage score
    ws_chart_data.append([])
    ws_chart_data.append(["Paper", "Score"])
    sorted_p = sorted(PAPERS, key=lambda x: sum(cov_score(v) for v in x["nc"]), reverse=True)[:12]
    for p in sorted_p:
        ws_chart_data.append([p["short_name"], sum(cov_score(v) for v in p["nc"])])

    chart2 = BarChart()
    chart2.type = "bar"
    chart2.title = "Top 12 Papers by NC Coverage Breadth"
    chart2.x_axis.title = "Coverage Score (Full=3, Partial=2, Taxonomy=1)"
    d2 = Reference(ws_chart_data, min_col=2, min_row=13, max_row=13 + len(sorted_p))
    c2 = Reference(ws_chart_data, min_col=1, min_row=14, max_row=13 + len(sorted_p))
    chart2.add_data(d2, titles_from_data=True)
    chart2.set_categories(c2)
    chart2.width = 16
    chart2.height = 12
    ws_m.add_chart(chart2, "N22")

    # --- Sheet 4: YP Gap Analysis ---
    ws_g = wb.create_sheet("YP_Gap_Analysis")
    g_headers = ["NC Category", "Sub-area", "YP Gap", "Fix / Source Paper", "Priority"]
    ws_g.append(g_headers)
    style_header(ws_g)
    pri_fill = {
        "Critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "High": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Medium": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Low": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "Low-Medium": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    }
    for g in YP_GAPS:
        ws_g.append(list(g))
    for r in range(2, ws_g.max_row + 1):
        pri = ws_g.cell(r, 5).value
        if pri in pri_fill:
            ws_g.cell(r, 5).fill = pri_fill[pri]
    auto_width(ws_g)

    # Pie chart: gaps by priority
    pri_counts = {}
    for g in YP_GAPS:
        pri_counts[g[4]] = pri_counts.get(g[4], 0) + 1
    start = ws_chart_data.max_row + 2
    ws_chart_data.cell(start, 1, "Priority")
    ws_chart_data.cell(start, 2, "Count")
    for i, (k, v) in enumerate(pri_counts.items(), 1):
        ws_chart_data.cell(start + i, 1, k)
        ws_chart_data.cell(start + i, 2, v)
    pie = PieChart()
    pie.title = "YP Gaps by Priority"
    pd = Reference(ws_chart_data, min_col=2, min_row=start, max_row=start + len(pri_counts))
    pc = Reference(ws_chart_data, min_col=1, min_row=start + 1, max_row=start + len(pri_counts))
    pie.add_data(pd, titles_from_data=True)
    pie.set_categories(pc)
    pie.width = 10
    pie.height = 8
    ws_g.add_chart(pie, "G2")

    # --- Sheet 5: Synthesis (network_context survey) ---
    ws_s = wb.create_sheet("Synthesis_Missing_Context")
    ws_s.append(["Missing Context", "Definition", "How to Extract", "Paper Backing"])
    style_header(ws_s)
    for row in SYNTHESIS_GAPS:
        ws_s.append(list(row))
    auto_width(ws_s)

    ws_s.append([])
    ws_s.append(["Professor Problem", "Root Cause (Feature Gap)", "Fix", "Papers"])
    style_header(ws_s, ws_s.max_row)
    for row in PROBLEM_MAPPING:
        ws_s.append(list(row))
    auto_width(ws_s)

    # --- Sheet 6: Violation Methods ---
    ws_v = wb.create_sheet("Violation_Detection")
    ws_v.append(["Paper", "Violation / Property Type", "Method", "Output on Violation"])
    style_header(ws_v)
    violations = [
        ("Yaksha-Prashna", "RAW/WAR/WAW field deps in chain", "Prolog queries on KB", "Dependency facts"),
        ("PREVAIL", "OOB, illegal pointer arithmetic", "Abstract interpretation", "Safety violation location"),
        ("Klint", "Spec vs ghost map op sequence", "KLEE + VeriFast", "Counterexample path"),
        ("Vigor", "Behavioral summary vs Python spec", "KLEE all paths", "Violating path"),
        ("NetSMC", "Bad state reachable", "Backward reachability SMC", "Witness trace"),
        ("Alpernas SAS'18", "Isolation / reachability", "Abstract interpretation fixpoint", "May have false positives"),
        ("BinPRE", "PRE accuracy", "Ground truth comparison", "F1 metrics"),
        ("NetLifter", "Protocol impl discrepancy", "AFG comparison", "CVE-level mismatch"),
        ("Aquila", "Undefined table behavior", "SMT SAT/UNSAT", "Bug localization slice"),
        ("Aquila", "Pipeline assertion violation", "∃ packet violating spec", "Culprit table/action"),
        ("Hydra", "LTL_f temporal property", "Runtime P4 monitor", "Packet capture witness"),
        ("VeriFlow", "Reachability, loop, isolation", "Equivalence class check", "Per-EC violation"),
        ("NetKAT", "Policy equivalence, isolation", "Equational reasoning", "Algebraic proof/counter"),
        ("Jones RG", "Guarantee ⊉ Rely in chain", "Formal R/G calculus", "Chain link mismatch"),
        ("HSA", "Slice leak, forwarding loop", "Header space intersection", "Leaking header space"),
        ("PIX/BOLT", "Performance contract breach", "Symbolic perf analysis", "Packet type + latency"),
    ]
    for v in violations:
        ws_v.append(list(v))
    auto_width(ws_v)

    # --- Sheet 7: Domain grouping (survey 1) ---
    ws_d = wb.create_sheet("Survey1_Domains")
    ws_d.append(["Domain", "Papers", "Focus"])
    style_header(ws_d)
    domains = [
        ("DOMAIN 1: eBPF-Specific", "YP, PREVAIL, eBPF Bit-Precise", "eBPF CFG analysis, memory regions"),
        ("DOMAIN 2: NF Binary Verification", "Klint, Vigor", "Ghost maps, behavioral summary"),
        ("DOMAIN 3: Stateful Network Verification", "NetSMC, Alpernas SAS'18, Alpernas FMCS'21", "State tables, packet effect, complexity"),
        ("DOMAIN 4: Protocol RE", "BinPRE, Protocol RE Survey", "Field roles, PRE phases"),
        ("DOMAIN 5: Production Data Plane", "Aquila, Alpernas (Cartesian)", "P4/SMT, sequential encoding"),
        ("DOMAIN 6: Binary Code Understanding", "Bin2Summary, CLAP", "Slicing, embeddings"),
    ]
    for d in domains:
        ws_d.append(list(d))
    auto_width(ws_d)

    # --- Sheet 8: Comparison YP vs literature ---
    ws_c = wb.create_sheet("YP_vs_Literature")
    ws_c.append(["Dimension", "Yaksha-Prashna", "Best Literature Alternative", "Gap Severity"])
    style_header(ws_c)
    comparisons = [
        ("Speed", "13–300 ms/NF", "Klint 2.7s–1.5min", "YP 200–1000× faster"),
        ("Map semantics", "map name + field", "Klint ghost map full schema", "Critical"),
        ("Field roles", "field name only", "BinPRE TYPE/LENGTH/SEQ", "High"),
        ("Protocol parse tree", "accessesProtocol only", "NetLifter AFG", "High"),
        ("Behavioral contract", "None", "Vigor/Klint Python spec", "Critical"),
        ("Temporal logic", "None (§7 limitation)", "NF-SE, Hydra, Patent", "High"),
        ("Chain map deps", "map name only", "NetSMC state evolution", "Critical"),
        ("Per-packet-type action", "per-path only", "PIX + Alpernas packet effect", "High"),
        ("Violation queries", "Prolog 24 predicates", "SMT (Aquila), SMC (NetSMC)", "Complementary"),
        ("eBPF-specific", "Yes (XDP only)", "PREVAIL, VEP, PIX on eBPF", "YP unique niche"),
    ]
    for row in comparisons:
        ws_c.append(list(row))
    auto_width(ws_c)

    # Bar chart: gap severity count
    ws_c.append([])
    ws_c.append(["Severity", "Count"])
    sev = {"Critical": 3, "High": 5, "Complementary": 1, "YP unique niche": 1}
    for k, v in [("Critical", 3), ("High", 5), ("Medium", 2)]:
        ws_c.append([k, v])

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")
    print(f"Papers: {len(PAPERS)} | Sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
